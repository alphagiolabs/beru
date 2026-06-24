"""Genera la plantilla Excel funcional para la feature 'Texto en lote'.

El feature toma filas de un Excel y las inyecta como texto en regiones
definidas sobre un video plantilla. Matching por ID == nombre de video
(sin extension, lowercase, trim). Una columna por region.

Salida: resources/plantilla_texto_en_lote.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = r"C:\Users\HIDROAA\Desktop\beru\resources\plantilla_texto_en_lote.xlsx"

ID_ALIASES = ["id", "code", "codigo", "video", "archivo",
              "filename", "name", "nombre", "identificador"]
N_REGIONS = 3
REGION_LABELS = [f"TEXT_{i+1}" for i in range(N_REGIONS)]

SAMPLES = [
    ("intro_001",      "OFERTA BLACK FRIDAY",  "50% OFF",        "Hasta 30/11"),
    ("promo_navidad",  "Feliz Navidad",        "Envios gratis",  "Solo esta semana"),
    ("salida_q4_2024", "Resumen del trimestre","$1.250.000",     "Crecimiento 12%"),
    ("tutorial_01",    "Como instalar Beru",   "Paso 1 de 5",    "www.beru.app"),
    ("intro_001",      "OFERTA BLACK FRIDAY",  "60% OFF",        "Hasta 01/12"),
    ("",               "fila sin id",          "no matchea",     "queda en reporte"),
]

THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", fgColor="1F2937")
HDR_FONT = Font(bold=True, color="FFFFFF")
ID_FILL = PatternFill("solid", fgColor="FEF3C7")
INPUT_FILL = PatternFill("solid", fgColor="ECFDF5")


def style_header(ws, row, cols):
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def build():
    wb = Workbook()

    # ---------- Instrucciones ----------
    ws_i = wb.active
    ws_i.title = "Instrucciones"
    ws_i["A1"] = "Plantilla: Texto en lote"
    ws_i["A1"].font = Font(size=16, bold=True)
    ws_i.merge_cells("A1:E1")

    rows = [
        ("Como funciona",
         "1) Define regiones de texto sobre un video (rectangulos con etiqueta TEXT_1, TEXT_2, ...). "
         "2) Importa este Excel desde el panel 'Texto en lote'. "
         "3) El sistema vincula cada fila con un video por su nombre (sin extension) y "
         "copia el contenido de cada columna a la region correspondiente."),
        ("Columna ID (obligatoria)",
         "Encabezado debe ser uno de: " + ", ".join(ID_ALIASES) +
         ". Valor = nombre del archivo de video SIN extension (ej. 'intro_001' para 'intro_001.mp4'). "
         "Comparacion: lowercase + trim en ambos lados."),
        ("Columnas de texto",
         f"Esta plantilla incluye {N_REGIONS} regiones: " + ", ".join(REGION_LABELS) +
         ". El encabezado debe coincidir (case-insensitive, trim) con la etiqueta de la region. "
         "Puedes renombrar/agregar columnas para que coincidan con tus regiones reales."),
        ("Lo que NO hace",
         "No transforma el texto (no hay UPPER, TRIM, etc.). El contenido se inyecta tal cual. "
         "Si necesitas normalizar, hazlo en otra columna del Excel antes."),
        ("Errores comunes",
         "ID vacio -> fila no matchea. ID duplicado -> marcada como 'dup' en el reporte. "
         "Encabezado de ID que no este en los aliases -> usar el selector manual del modal de mapeo. "
         "Columna de region con nombre distinto a la etiqueta -> usar el selector manual."),
        ("Validacion",
         "La hoja 'Validacion' muestra el ID normalizado y avisa duplicados/vacios con formulas."),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        ws_i.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws_i.cell(row=i, column=2, value=v)
        ws_i.merge_cells(start_row=i, start_column=2, end_row=i, end_column=5)
        ws_i.row_dimensions[i].height = 50
        ws_i.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws_i.column_dimensions["A"].width = 22
    for c in "BCDE":
        ws_i.column_dimensions[c].width = 25
    ws_i.row_dimensions[1].height = 24

    # ---------- Datos (hoja que se importa) ----------
    ws_d = wb.create_sheet("Datos")
    headers = ["id"] + REGION_LABELS
    ws_d.append(headers)
    style_header(ws_d, 1, list(range(1, len(headers) + 1)))

    for i, row in enumerate(SAMPLES, start=2):
        for j, val in enumerate(row, start=1):
            cell = ws_d.cell(row=i, column=j, value=val)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.fill = ID_FILL if j == 1 else INPUT_FILL
    ws_d.column_dimensions["A"].width = 24
    for i in range(N_REGIONS):
        ws_d.column_dimensions[get_column_letter(2 + i)].width = 28
    ws_d.freeze_panes = "B2"

    # Dropdown para el encabezado de la columna A (aliases)
    dv_id = DataValidation(type="list", formula1='"' + ",".join(ID_ALIASES) + '"', allow_blank=False)
    dv_id.add("A1")
    ws_d.add_data_validation(dv_id)

    # ---------- Validacion (formulas de chequeo) ----------
    ws_v = wb.create_sheet("Validacion")
    v_headers = ["id_original", "id_normalizado", "estado", "veces", "TEXT_1_recuento"]
    ws_v.append(v_headers)
    style_header(ws_v, 1, list(range(1, len(v_headers) + 1)))

    n = len(SAMPLES)
    for i in range(n):
        r = i + 2
        ws_v.cell(row=r, column=1, value=f"=Datos!A{r}").border = BORDER
        ws_v.cell(row=r, column=2,
                  value=f'=IF(TRIM(Datos!A{r})="","",LOWER(TRIM(Datos!A{r})))'
                  ).border = BORDER
        # estado: vacio | duplicado | ok
        ws_v.cell(row=r, column=3,
                  value=(f'=IF(TRIM(Datos!A{r})="","VACIO",'
                         f'IF(COUNTIF(Datos!A:A,Datos!A{r})>1,"DUPLICADO","OK"))')
                  ).border = BORDER
        # veces que aparece
        ws_v.cell(row=r, column=4,
                  value=f'=IF(TRIM(Datos!A{r})="",0,COUNTIF(Datos!A:A,Datos!A{r}))'
                  ).border = BORDER
        # longitud del texto en TEXT_1 como sanity check
        ws_v.cell(row=r, column=5,
                  value=f'=LEN(Datos!B{r})'
                  ).border = BORDER
    # formato condicional via relleno segun estado (no usamos ConditionalFormatting
    # para mantenerlo simple y compatible con todas las versiones)
    for i in range(n):
        r = i + 2
        ws_v.cell(row=r, column=1).fill = INPUT_FILL
        ws_v.cell(row=r, column=2).fill = INPUT_FILL
        ws_v.cell(row=r, column=4).fill = INPUT_FILL
        ws_v.cell(row=r, column=5).fill = INPUT_FILL

    ws_v.column_dimensions["A"].width = 24
    ws_v.column_dimensions["B"].width = 24
    ws_v.column_dimensions["C"].width = 14
    ws_v.column_dimensions["D"].width = 10
    ws_v.column_dimensions["E"].width = 16
    ws_v.freeze_panes = "A2"

    # Resumen arriba (KPIs)
    ws_v.insert_rows(1)
    ws_v.cell(row=1, column=1, value="Resumen").font = Font(bold=True)
    ws_v.cell(row=1, column=2,
              value=(f'="Filas: "&COUNTA(Datos!A2:A{n+1})&"  |  OK: "'
                     f'&COUNTIF(C3:C{n+1},"OK")&"  |  Duplicados: "'
                     f'&COUNTIF(C3:C{n+1},"DUPLICADO")&"  |  Vacios: "'
                     f'&COUNTIF(C3:C{n+1},"VACIO")'))
    ws_v.merge_cells(start_row=1, start_column=2, end_row=1, end_column=5)
    ws_v.row_dimensions[1].height = 20
    # re-aplicar estilo de header en la fila 2 (que ahora es la de encabezados real)
    style_header(ws_v, 2, [1, 2, 3, 4, 5])

    wb.save(OUT)
    print(f"OK -> {OUT}")


if __name__ == "__main__":
    build()
